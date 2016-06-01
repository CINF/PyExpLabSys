#!/usr/bin/env python3

"""File parser for the Avantage xlsx export format"""

from __future__ import print_function, unicode_literals

import logging
log = logging.getLogger(__name__)
log.addHandler(logging.NullHandler())

import numpy
import openpyxl


class AvantageXLSXExport(dict):
    """Class that implements a parser for the Avantage xldx export format

    The class inherits form dict, so the components can be accessed in the normal dict
    fasion i.e:

    >>> avexport = AvantageXLSXExport("...path...")
    >>> avexport.keys()
    ['Ni 2p Scan', 'Pt 4f Scan']
    spectrum = avexport['Ni 2p Scan']

    """

    def __init__(self, filepath):
        log.info("Opened workbook %s", filepath)
        wb = openpyxl.load_workbook(filepath)

        # Iterate over sheets and parse depending on type
        for sheet in wb.worksheets:
            # Check if it is a spectrum
            if sheet["H1"].value == "Acquisition Parameters :":
                self[sheet.title] = Spectrum(sheet)
                continue

            # Check if it is a table summary
            if sheet["A1"].value == "Peak Table : ":
                self[sheet.title] = PeakTable(sheet)
                continue

            log.debug("Type if sheet %s is unknown, not parsed", sheet)


class Spectrum(dict):
    """Class that represents a spectrum sheet"""

    def __init__(self, sheet):
        """Parse spectrum from sheet"""
        log.debug("Parse sheet %s as spectrum", sheet)
        number_of_elements = sheet["D7"].value
        for column in sheet.columns:
            # Get the cell that contains the unit and skip the column if it is empty
            unit_cell = column[15]
            if unit_cell.value is None:
                continue

            # Get the name and calculate start and end coordinates
            name = unit_cell.offset(row=-1).value or "Counts"
            start = unit_cell.offset(row=1).coordinate
            end = unit_cell.offset(row=number_of_elements).coordinate
            log.debug("Parsing data: %s in cell %s:%s", name, start, end)

            # Parse the data without intermediate copies
            row_iter = sheet.iter_rows(start + ":" + end)  # Iterator over rows
            value_generator = (row[0].value for row in row_iter)  # Generator over values
            array = numpy.fromiter(value_generator, dtype=float, count=number_of_elements)

            self[name] = array


class PeakTable(dict):
    """Class that represents a peak table"""

    def __init__(self, sheet):
        """Parse peak table from sheet"""
        log.debug("Parse peak table from sheet %s", sheet)

        for row in sheet:
            # If a table is about to start
            if row[0].value is not None and row[0].value.startswith('Peak'):
                table_name = row[0].value.strip(" :").lower()

                # Calculate start and end cell. Assume start cell is below and if not
                # below and 1 to the right
                start = row[0].offset(row=1)
                if start.value is None or start.value == " ":
                    start = start.offset(column=1)

                # Move to the right and down to find end cell
                end = start
                while end.offset(column=1).value is not None:
                     end = end.offset(column=1)
                while end.offset(row=1).value is not None:
                     end = end.offset(row=1)
                log.debug('Found table "%s" %s:%s'.format(table_name, start.coordinate,
                          end.coordinate))

                # Form the range string and calculate rows and columns
                range_str = start.coordinate + ":" + end.coordinate
                columns = end.col_idx - start.col_idx + 1
                rows = end.row - start.row + 1

                # Fill values into numpy object array
                array = numpy.empty(columns * rows, dtype=object).reshape(rows, columns)
                for row_num, row in enumerate(sheet.iter_rows(range_str)):
                    for col_num, cell in enumerate(row):
                        array[row_num, col_num] = cell.value

                self[table_name] = array
                
        
        


def test():
    log.setLevel(logging.DEBUG)
    log.addHandler(logging.StreamHandler())
    #logging.basicConfig()
    path = '/home/kenni/Dokumenter/xps/soren_dahl_battery/soren_dahl_battery/HT1/peak table.xlsx'
    avexport = AvantageXLSXExport(path)
    print(avexport["Ni2p Scan"].keys())


if __name__ == "__main__":
    test()
